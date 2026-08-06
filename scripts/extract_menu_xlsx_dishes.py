#!/usr/bin/env python3
"""Trich xuat "bua an" (Sang/Trua/Toi) that tu file thuc don noi bo du an
`data/Bang xac dinh nhu cau dinh duong + thuc don.xlsx` thanh dishes.csv +
dish_ingredients.csv bo sung.

Ticket: DAT-12 (mo rong mon Viet). Day KHONG phai LLM sang tac - la ETL tu
file Excel that trong data/, nhung viec KHOP TEN nguyen lieu -> food_id la
tu dong (fuzzy theo ten chuan hoa), nen van gan verified_by=pending de R2
kiem tra lai khop dung khong truoc khi tin dung cho benh nhan (RULE-2/3).

Cau truc file: moi sheet (td1, td2, "TD 3+4", "TD 5+6+7") la 1 chuoi
"Thuc don N" -> tung bua (Sang/Trua/Toi/Ca ngay) -> danh sach nguyen lieu
voi cot "KL song sach" (gram thuc te). Dong nhan "Hien tai"/"Can xay dung"
la dong tong ket, khong phai nguyen lieu.

Chay: python scripts/extract_menu_xlsx_dishes.py
"""

from __future__ import annotations

import csv
import glob
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
SEEDS = ROOT / "data" / "seeds"

END_MARKERS = {"Hiện tại", "Cần xây dựng", "Cần xây dựng cả ngày"}
MEAL_LABELS = {"Sáng", "Trưa", "Tối", "Cả ngày", "Xế", "Phụ"}
SHEETS = ["tđ1", "tđ2", "TĐ 3+4", "TĐ 5+6+7"]

DISH_HEADER = ["dish_id", "name_vi", "region", "serving_g", "verified_by", "note"]
ING_HEADER = ["dish_id", "food_id", "grams", "note"]


def _norm(name: str) -> str:
    name = unicodedata.normalize("NFC", name).strip().lower()
    name = re.sub(r"\([^)]*\)", "", name)  # bỏ chú thích trong ngoặc
    name = re.sub(r"\s+", " ", name).strip()
    return name


def _load_food_index() -> dict[str, str]:
    idx: dict[str, str] = {}
    with open(SEEDS / "food_items.csv", newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if not (row.get("kcal_100g") or "").strip():
                continue
            key = _norm(row["name_vi"])
            if key and key not in idx:
                idx[key] = row["id"]
    return idx


def _known_dish_ids() -> set[str]:
    ids: set[str] = set()
    with open(SEEDS / "dishes.csv", newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ids.add(row["dish_id"])
    return ids


def extract_sheet(ws, sheet_name: str, food_idx: dict[str, str]) -> tuple[list[dict], list[dict], dict]:
    dishes: list[dict] = []
    ings: list[dict] = []
    stats = {"meals_found": 0, "meals_kept": 0, "ingredients_unmatched": 0, "ingredients_bad_grams": 0}

    thuc_don = None
    meal_label = None
    meal_idx = 0
    cur_ings: list[tuple[str, float, str]] = []

    def flush():
        nonlocal cur_ings
        if meal_label and cur_ings:
            stats["meals_found"] += 1
            dish_id = f"MENU-{sheet_name}-TD{thuc_don}-{meal_label}-{meal_idx}".replace(" ", "")
            total_g = sum(g for _, g, _ in cur_ings)
            dishes.append(
                {
                    "dish_id": dish_id,
                    "name_vi": f"Bữa {meal_label.lower()} - Thực đơn {thuc_don} ({sheet_name})",
                    "region": "",
                    "serving_g": f"{total_g:g}",
                    "verified_by": "pending",
                    "note": (
                        f"Trích từ data/Bảng xác định nhu cầu dinh dưỡng + thực đơn.xlsx, "
                        f"sheet '{sheet_name}', Thực đơn {thuc_don}, bữa {meal_label}. "
                        "Khớp food_id tự động theo tên chuẩn hoá — CẦN R2 rà soát trước khi dùng."
                    ),
                }
            )
            for fid, g, name in cur_ings:
                ings.append({"dish_id": dish_id, "food_id": fid, "grams": f"{g:g}", "note": name})
            stats["meals_kept"] += 1
        cur_ings = []

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col3 = row[3] if len(row) > 3 else None

        if col1.startswith("Thực đơn"):
            flush()
            m = re.search(r"\d+", col1)
            thuc_don = m.group(0) if m else "?"
            meal_label = None
            continue

        if col1 in END_MARKERS or col1 == "Tên TP":
            flush()
            meal_label = None
            continue

        if col0 in MEAL_LABELS:
            flush()
            meal_label = col0
            meal_idx += 1
            # dòng này CŨNG là 1 nguyên liệu (col1 = tên NL, không phải tên món)
            col0 = ""  # rơi xuống nhánh nguyên liệu bên dưới

        if not meal_label:
            continue
        if not col1 or col1 in END_MARKERS:
            continue

        key = _norm(col1)
        fid = food_idx.get(key)
        try:
            grams = float(col3) if col3 is not None else None
        except (ValueError, TypeError):
            grams = None

        if grams is None or not (0 < grams <= 2000):
            stats["ingredients_bad_grams"] += 1
            continue
        if fid is None:
            stats["ingredients_unmatched"] += 1
            continue
        cur_ings.append((fid, grams, col1))

    flush()
    return dishes, ings, stats


def main() -> int:
    files = [f for f in glob.glob(str(ROOT / "data" / "*.xlsx")) if "PURINE" not in f]
    if not files:
        print("Không tìm thấy file thực đơn xlsx trong data/")
        return 1
    wb = openpyxl.load_workbook(files[0], data_only=True)

    food_idx = _load_food_index()
    known_dish_ids = _known_dish_ids()

    all_dishes: list[dict] = []
    all_ings: list[dict] = []
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"Bỏ qua sheet không tồn tại: {sheet_name}")
            continue
        ws = wb[sheet_name]
        dishes, ings, stats = extract_sheet(ws, sheet_name, food_idx)
        print(
            f"[{sheet_name}] bữa tìm thấy={stats['meals_found']} giữ={stats['meals_kept']} "
            f"NL không khớp food_id={stats['ingredients_unmatched']} NL gram bất thường={stats['ingredients_bad_grams']}"
        )
        for d in dishes:
            if d["dish_id"] not in known_dish_ids:
                all_dishes.append(d)
        all_ings.extend([i for i in ings if i["dish_id"] in {d["dish_id"] for d in dishes}])

    dish_out = SEEDS / "dishes.menu_xlsx.csv"
    ing_out = SEEDS / "dish_ingredients.menu_xlsx.csv"
    with open(dish_out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=DISH_HEADER)
        w.writeheader()
        w.writerows(all_dishes)
    with open(ing_out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=ING_HEADER)
        w.writeheader()
        w.writerows(all_ings)

    print(f"\nTổng: {len(all_dishes)} bữa/món, {len(all_ings)} dòng nguyên liệu")
    print(f"Đã ghi {dish_out}, {ing_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
