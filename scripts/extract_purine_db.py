#!/usr/bin/env python3
"""Trích bảng purine từ `data/PURINEDATABASEANDDATASOURCES2025.xlsx` thành CSV
tham chiếu phẳng, kèm trích dẫn nguồn gốc từng dòng.

Ticket: DAT-14. LLM: NO — thuần ETL, mọi con số đi thẳng từ file nguồn.

NGUỒN: "USDA and ODS-NIH Database for the Purine Content of Common Foods"
(Release 2.0, 2025). File gồm:
  - Table1_Food data_NAm sources  — nguồn Bắc Mỹ
  - Table2_Food data_nonNAm sources — nguồn ngoài Bắc Mỹ (Nhật, Trung Quốc,
    Hàn Quốc...). Nhóm này QUAN TRỌNG với dự án vì thực phẩm châu Á (cá, hải
    sản, nội tạng, đậu nành) gần với thực đơn Việt Nam hơn dữ liệu Bắc Mỹ.
  - Table3_Alcohol data — đồ uống có cồn
  - Table6_Sources of data — chú giải mã nguồn (a, b, c...) → trích dẫn đầy đủ

Cột `purine_mg` của dự án dùng **"Total of 4 Purine Bases"** (adenine +
guanine + hypoxanthine + xanthine, mg/100g) — đúng đại lượng mà hướng dẫn
gout dùng để xếp nhóm thực phẩm. KHÔNG dùng cột "Uric Acid".

⚠️ Đây là bảng THAM CHIẾU, chưa phải `purine_values.csv`. Việc khớp món Việt
với mô tả tiếng Anh là bước riêng, phải ghi rõ dòng nào là proxy nước ngoài
(RULE-2/DEC-008) — xem `scripts/map_purine_to_food_items.py`.

Chạy: python scripts/extract_purine_db.py
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "PURINEDATABASEANDDATASOURCES2025.xlsx"
OUT = ROOT / "data" / "seeds" / "purine_db_reference.csv"

SHEETS = {
    "Table1_Food data_NAm sources": "NAm",
    "Table2_Food data_nonNAm sources": "nonNAm",
    "Table3_Alcohol data": "alcohol",
}

COL_FOOD = 0
COL_N = 1
COL_TOTAL_PURINE = 18
COL_YEAR = 20
COL_SOURCE_CODE = 21
COL_COUNTRY = 22
DATA_START_ROW = 6  # 2 dòng tiêu đề gộp + tiêu đề bảng


def load_source_citations(wb: openpyxl.Workbook) -> dict[str, str]:
    """Mã nguồn (a/b/c...) -> trích dẫn đầy đủ, từ Table6."""
    ws = wb["Table6_Sources of data"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, citation, year, country = row[0], row[1], row[2], row[3]
        if not code or not citation:
            continue
        out[str(code).strip()] = f"{str(citation).strip()} [{country}, {year}]"
    return out


def _num(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in {"", "-", "nd", "ND", "NA"}:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def main() -> None:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    citations = load_source_citations(wb)
    print(f"Table6: {len(citations)} ma nguon")

    rows: list[dict[str, object]] = []
    for sheet, group in SHEETS.items():
        ws = wb[sheet]
        kept = 0
        for rec in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            name = rec[COL_FOOD]
            if not name or not str(name).strip():
                continue
            total = _num(rec[COL_TOTAL_PURINE])
            if total is None:
                # Dòng tiêu đề nhóm (VD "Beef Organ Products") — không có số
                continue
            code = str(rec[COL_SOURCE_CODE] or "").strip()
            rows.append(
                {
                    "food_description_en": str(name).strip(),
                    "total_purine_mg_100g": total,
                    "n_samples": str(rec[COL_N] or "").strip(),
                    "country_of_origin": str(rec[COL_COUNTRY] or "").strip(),
                    "year": str(rec[COL_YEAR] or "").strip(),
                    "source_group": group,
                    "source_code": code,
                    "source_citation": citations.get(code, ""),
                }
            )
            kept += 1
        print(f"  {sheet}: {kept} dong co so purine")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "food_description_en",
                "total_purine_mg_100g",
                "n_samples",
                "country_of_origin",
                "year",
                "source_group",
                "source_code",
                "source_citation",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    no_citation = sum(1 for r in rows if not r["source_citation"])
    print(f"\nDa ghi {OUT.relative_to(ROOT)}: {len(rows)} dong")
    if no_citation:
        print(f"CANH BAO: {no_citation} dong khong map duoc trich dan Table6")


if __name__ == "__main__":
    main()
