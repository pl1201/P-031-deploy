#!/usr/bin/env python3
"""Trich xuat sheet "Bang TP co phospho" trong file thuc don noi bo du an
`data/Bang xac dinh nhu cau dinh duong + thuc don.xlsx` thanh cac dong moi
cho `food_items.csv`.

Ticket: DAT-04/DAT-12. Nguon: bang thanh phan dinh duong goc tu NIN, do cac
chuyen gia dinh duong cua du an dang su dung (xac nhan qua chat, khong phai
LLM sang tac, khong phai suy doan).

Ly do dung sheet "Bang TP co phospho" (397 dong) thay vi "Bang TP" (841 dong,
nhieu ten hon nhung KHONG co Na/K/Phospho): validate_data.py yeu cau
na_mg/k_mg/p_mg bat buoc phai co gia tri khi kcal_100g da dien (khong nam
trong OPTIONAL_NUMERIC_COLS) - "Bang TP" thieu 3 cot nay nen khong the dung
de tao dong moi ma khong vi pham RULE-2/R40.3.

Chinh sach xung dot (DEC-008 "de trong con hon tich hop sai"):
- Neu ten (chuan hoa) da co san trong food_items.csv: BO QUA hoan toan, khong
  ghi de gia tri da duoc nguoi truoc xac minh, du co lech so voi bang xlsx nay
  (da phat hien ~21/80 dong lech kcal >2 khi doi chieu thu cong - co the do
  khac an ban NIN hoac lam tron cua nguoi bien soan bang noi bo).
- Ten trung lap NOI BO trong chinh sheet nay (vd "Dau ha lan" xuat hien 2 dong
  gia tri khac nhau): giu dong xuat hien dau tien, bo qua cac dong sau.

Chay: python scripts/extract_menu_xlsx_composition.py [--out PATH] [--start-id N]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
SEEDS = ROOT / "data" / "seeds"
XLSX = ROOT / "data" / "Bảng xác định nhu cầu dinh dưỡng + thực đơn.xlsx"
SHEET = "Bảng TP có phospho"

SOURCE_REF = (
    "NIN - Bảng thành phần dinh dưỡng (bảng nội bộ do chuyên gia dinh dưỡng "
    "dự án biên soạn/sử dụng từ nguồn NIN, file "
    "'Bảng xác định nhu cầu dinh dưỡng + thực đơn.xlsx', sheet 'Bảng TP có phospho')"
)

HEADER = [
    "id",
    "name_vi",
    "aliases",
    "category",
    "kcal_100g",
    "protein_g",
    "carb_g",
    "fat_g",
    "fiber_g",
    "sugar_g",
    "na_mg",
    "k_mg",
    "p_mg",
    "purine_mg",
    "purine_source_ref",
    "gi_index",
    "gi_source",
    "gi_source_ref",
    "contains_allergens",
    "source",
    "source_ref",
    "is_estimated",
]

# Cot trong sheet (1-indexed), theo header thuc te da doc bang openpyxl.
COL_NAME = 1
COL_KCAL = 3  # E/100
COL_PROTEIN = 5  # P/100
COL_FAT = 7  # L/100
COL_CARB = 9  # G/100
COL_FIBER = 11  # Cellu/100
COL_NA = 13  # Na/100
COL_K = 15  # K/100
COL_P = 21  # Phos/100


def _norm(name: str) -> str:
    name = unicodedata.normalize("NFC", name).strip().lower()
    name = re.sub(r"\([^)]*\)", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def _num(v: object) -> str:
    if v is None:
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ""
    return f"{f:.2f}".rstrip("0").rstrip(".") if "." in f"{f:.2f}" else f"{f:.2f}"


def load_existing_names() -> set[str]:
    names: set[str] = set()
    with open(SEEDS / "food_items.csv", newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            names.add(_norm(row["name_vi"]))
    return names


def extract(start_id: int) -> tuple[list[dict], dict]:
    existing = load_existing_names()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]

    seen_this_sheet: set[str] = set()
    rows: list[dict] = []
    stats = {"total": 0, "skipped_existing": 0, "skipped_dup": 0, "skipped_incomplete": 0, "added": 0}
    next_id = start_id

    for r in range(3, ws.max_row + 1):
        raw_name = ws.cell(r, COL_NAME).value
        if not raw_name or not str(raw_name).strip():
            continue
        stats["total"] += 1
        name = str(raw_name).strip()
        key = _norm(name)

        if key in existing:
            stats["skipped_existing"] += 1
            continue
        if key in seen_this_sheet:
            stats["skipped_dup"] += 1
            continue

        kcal = _num(ws.cell(r, COL_KCAL).value)
        protein = _num(ws.cell(r, COL_PROTEIN).value)
        fat = _num(ws.cell(r, COL_FAT).value)
        carb = _num(ws.cell(r, COL_CARB).value)
        fiber = _num(ws.cell(r, COL_FIBER).value)
        na = _num(ws.cell(r, COL_NA).value)
        k = _num(ws.cell(r, COL_K).value)
        p = _num(ws.cell(r, COL_P).value)

        if not all([kcal, protein, fat, carb, fiber, na, k, p]):
            stats["skipped_incomplete"] += 1
            continue

        seen_this_sheet.add(key)
        rows.append(
            {
                "id": str(next_id),
                "name_vi": name,
                "aliases": "",
                "category": "",
                "kcal_100g": kcal,
                "protein_g": protein,
                "carb_g": carb,
                "fat_g": fat,
                "fiber_g": fiber,
                "sugar_g": "",
                "na_mg": na,
                "k_mg": k,
                "p_mg": p,
                "purine_mg": "",
                "purine_source_ref": "",
                "gi_index": "",
                "gi_source": "",
                "gi_source_ref": "",
                "contains_allergens": "",
                "source": "NIN",
                "source_ref": SOURCE_REF,
                "is_estimated": "FALSE",
            }
        )
        next_id += 1
        stats["added"] += 1

    return rows, stats


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", type=Path, default=SEEDS / "food_items.csv")
    ap.add_argument("--start-id", type=int, default=3000, help="ID bắt đầu cấp cho dòng mới")
    args = ap.parse_args()

    rows, stats = extract(args.start_id)

    with open(args.out, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADER)
        for row in rows:
            writer.writerow(row)

    print(f"Tổng dòng đọc từ sheet: {stats['total']}")
    print(f"Bỏ qua (đã có trong food_items.csv): {stats['skipped_existing']}")
    print(f"Bỏ qua (trùng tên nội bộ trong sheet): {stats['skipped_dup']}")
    print(f"Bỏ qua (thiếu 1 trong các cột bắt buộc): {stats['skipped_incomplete']}")
    print(f"Đã thêm: {stats['added']} dòng mới, id {args.start_id}..{args.start_id + stats['added'] - 1}")


if __name__ == "__main__":
    main()
